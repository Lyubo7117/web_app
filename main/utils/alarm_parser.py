# -*- coding: utf-8 -*-
"""预警数据解析模块 - 修复城市名空白（支持从预警内容提取地级市）"""

import pandas as pd
import glob
import os
import re
from openpyxl import load_workbook


# ══════════════════════════════════════
# 中国地级行政区名称库（用于从预警文本中提取）
# 覆盖：333个地级市 + 自治州 + 地区 + 盟
# ══════════════════════════════════════
_PREFECTURE_CITIES = sorted([
    # 直辖市
    '北京', '天津', '上海', '重庆',
    # 河北（11）
    '石家庄', '唐山', '秦皇岛', '邯郸', '邢台', '保定',
    '张家口', '承德', '沧州', '廊坊', '衡水',
    # 山西（11）
    '太原', '大同', '阳泉', '长治', '晋城', '朔州',
    '晋中', '运城', '忻州', '临汾', '吕梁',
    # 内蒙古（9盟市）
    '呼和浩特', '包头', '乌海', '赤峰', '通辽',
    '鄂尔多斯', '呼伦贝尔', '巴彦淖尔', '乌兰察布',
    '兴安', '锡林郭勒', '阿拉善',
    # 辽宁（14）
    '沈阳', '大连', '鞍山', '抚顺', '本溪', '丹东',
    '锦州', '营口', '阜新', '辽阳', '盘锦', '铁岭',
    '朝阳', '葫芦岛',
    # 吉林（8+1自治州）
    '长春', '吉林', '四平', '辽源', '通化', '白山',
    '松原', '白城', '延边',
    # 黑龙江（12+1地区）
    '哈尔滨', '齐齐哈尔', '鸡西', '鹤岗', '双鸭山', '大庆',
    '伊春', '佳木斯', '七台河', '牡丹江', '黑河', '绥化',
    '大兴安岭',
    # 江苏（13）
    '南京', '无锡', '徐州', '常州', '苏州', '南通',
    '连云港', '淮安', '盐城', '扬州', '镇江', '泰州', '宿迁',
    # 浙江（11）
    '杭州', '宁波', '温州', '嘉兴', '湖州', '绍兴',
    '金华', '衢州', '舟山', '台州', '丽水',
    # 安徽（16）
    '合肥', '芜湖', '蚌埠', '淮南', '马鞍山', '淮北',
    '铜陵', '安庆', '黄山', '滁州', '阜阳', '宿州',
    '六安', '亳州', '池州', '宣城',
    # 福建（9）
    '福州', '厦门', '莆田', '三明', '泉州', '漳州',
    '南平', '龙岩', '宁德',
    # 江西（11）
    '南昌', '景德镇', '萍乡', '九江', '新余', '鹰潭',
    '赣州', '吉安', '宜春', '抚州', '上饶',
    # 山东（16）
    '济南', '青岛', '淄博', '枣庄', '东营', '烟台',
    '潍坊', '济宁', '泰安', '威海', '日照', '临沂',
    '德州', '聊城', '滨州', '菏泽',
    # 河南（17）
    '郑州', '开封', '洛阳', '平顶山', '安阳', '鹤壁',
    '新乡', '焦作', '濮阳', '许昌', '漯河', '三门峡',
    '南阳', '商丘', '信阳', '周口', '驻马店',
    # 湖北（12+1自治州）
    '武汉', '黄石', '十堰', '宜昌', '襄阳', '鄂州',
    '荆门', '孝感', '荆州', '黄冈', '咸宁', '随州',
    '恩施',
    # 湖南（13+1自治州）
    '长沙', '株洲', '湘潭', '衡阳', '邵阳', '岳阳',
    '常德', '张家界', '益阳', '郴州', '永州', '怀化', '娄底',
    '湘西',
    # 广东（21）
    '广州', '韶关', '深圳', '珠海', '汕头', '佛山',
    '江门', '湛江', '茂名', '肇庆', '惠州', '梅州',
    '汕尾', '河源', '阳江', '清远', '东莞', '中山',
    '潮州', '揭阳', '云浮',
    # 广西（14）
    '南宁', '柳州', '桂林', '梧州', '北海', '防城港',
    '钦州', '贵港', '玉林', '百色', '贺州', '河池',
    '来宾', '崇左',
    # 海南（4）
    '海口', '三亚', '三沙', '儋州',
    # 四川（18+3自治州）
    '成都', '自贡', '攀枝花', '泸州', '德阳', '绵阳',
    '广元', '遂宁', '内江', '乐山', '南充', '眉山',
    '宜宾', '广安', '达州', '雅安', '巴中', '资阳',
    '阿坝', '甘孜', '凉山',
    # 贵州（6+3自治州）
    '贵阳', '六盘水', '遵义', '安顺', '毕节', '铜仁',
    '黔西南', '黔东南', '黔南',
    # 云南（8+8自治州）
    '昆明', '曲靖', '玉溪', '保山', '昭通', '丽江',
    '普洱', '临沧',
    '楚雄', '红河', '文山', '西双版纳', '大理', '德宏',
    '怒江', '迪庆',
    # 西藏（6+1地区）
    '拉萨', '日喀则', '昌都', '林芝', '山南', '那曲',
    '阿里',
    # 陕西（10）
    '西安', '铜川', '宝鸡', '咸阳', '渭南', '延安',
    '汉中', '榆林', '安康', '商洛',
    # 甘肃（12+2自治州）
    '兰州', '嘉峪关', '金昌', '白银', '天水', '武威',
    '张掖', '平凉', '酒泉', '庆阳', '定西', '陇南',
    '甘南', '临夏',
    # 青海（2+6自治州）
    '西宁', '海东',
    '海北', '黄南', '海南州', '果洛', '玉树', '海西',
    # 宁夏（5）
    '银川', '石嘴山', '吴忠', '固原', '中卫',
    # 新疆（4+5自治州+5地区）
    '乌鲁木齐', '克拉玛依', '吐鲁番', '哈密',
    '昌吉', '博尔塔拉', '巴音郭楞', '阿克苏', '克孜勒苏',
    '喀什', '和田', '伊犁', '塔城', '阿勒泰',
    # 特别行政区
    '香港', '澳门',
], key=len, reverse=True)  # 按长度降序排列，优先匹配长名字


def _extract_city_from_text(text):
    """
    从预警内容文本中提取地级市/自治州名称。
    
    匹配策略（按优先级）：
    1. 精确匹配 _PREFECTURE_CITIES 中的已知城市名
    2. 正则匹配 "XX省XX市" / "XX省XX自治州" 中的地级名
    
    Parameters
    ----------
    text : str
        预警内容全文（如 "福建省泉州市南安市气象台发布..."）
    
    Returns
    -------
    str | None
        提取到的城市名，未找到返回 None
    """
    if not text or pd.isna(text):
        return None
    text = str(text).strip()
    if len(text) < 4:
        return None

    # ── 策略1: 在已知城市库中精确查找（按长度降序，避免短名误匹配）──
    for city in _PREFECTURE_CITIES:
        idx = text.find(city)
        if idx == -1:
            continue
        # 排除匹配到省份前缀的情况（如 "吉林省延边..." 不应匹配 "吉林"）
        # 如果匹配位置在开头或紧接"省/自治区"后面，且匹配到的名字恰好是已知省名 → 跳过
        province_prefixes = ['北京', '天津', '上海', '重庆', '河北', '山西', '内蒙古',
                             '辽宁', '吉林', '黑龙江', '江苏', '浙江', '安徽', '福建',
                             '江西', '山东', '河南', '湖北', '湖南', '广东', '广西',
                             '海南', '四川', '贵州', '云南', '西藏', '陕西', '甘肃',
                             '青海', '宁夏', '新疆']
        if idx <= 2 and city in province_prefixes:
            continue  # 跳过开头的省份名，继续找后面的地级市
        return city

    # ── 策略2: 正则提取 "XX省(XX市|XX自治州|XX地区|XX盟)" ──
    # 匹配如：贵州省黔西南布依族苗族自治州、吉林省延边朝鲜族自治州
    patterns = [
        r'省([^省市自治区]{2,15}(?:自治州))',
        r'省([^省市自治区]{2,10}市)',
        r'自治区([^省市自治区]{2,10}市)',
        r'自治区([^省市自治区]{2,15}(?:自治州))',
    ]
    for pat in patterns:
        m = re.search(pat, text)
        if m:
            name = m.group(1).strip()
            if name and len(name) >= 2:
                return name

    return None


def get_latest_alarms(data_folder=None):
    debug = []
    debug.append("[开始] get_latest_alarms 执行")

    base_dir = os.path.join(os.getcwd(), 'main', 'data_output', 'alarms')
    fallback_dir = os.path.join(os.path.dirname(__file__), '..', 'data_output', 'alarms')
    alarm_dir = base_dir if os.path.exists(base_dir) else fallback_dir

    if not os.path.exists(alarm_dir):
        debug.append(f"[错误] 目录不存在: {alarm_dir}")
        return pd.DataFrame(), None, debug

    pattern = os.path.join(alarm_dir, '全国预警信息_*.xlsx')
    files = glob.glob(pattern)
    # 备用递归搜索
    if not files:
        files = glob.glob(os.path.join(alarm_dir, '**', '全国预警信息_*.xlsx'), recursive=True)
        files = [f for f in files if not os.path.basename(f).startswith('~')]

    if not files:
        debug.append("[警告] 未找到预警Excel文件")
        return pd.DataFrame(), None, debug

    latest_file = max(files, key=os.path.getmtime)
    debug.append(f"[文件] 最新文件: {os.path.basename(latest_file)}")

    try:
        wb = load_workbook(latest_file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            debug.append("[错误] Excel 行数不足4行")
            return pd.DataFrame(), latest_file, debug

        # 第三行是列名，第四行起是数据
        header = [str(c).strip() if c else '' for c in rows[2]]
        data_rows = rows[3:]
        df = pd.DataFrame(data_rows, columns=header)
        df = df.dropna(how='all')
        debug.append(f"[解析] 原始行数: {len(df)}, 列数: {len(df.columns)}, 列名前8: {list(df.columns[:8])}")

        # 清理列名
        def clean_col(col):
            return re.sub(r'[\s\n\r]+', '', str(col))
        df.columns = [clean_col(c) for c in df.columns]

        required = ['省份', '城市', '预警类型', '预警等级', '发布时间', '解除时间']
        col_map = {}
        for req in required:
            for col in df.columns:
                if req in col:
                    col_map[req] = col
                    break

        if len(col_map) != len(required):
            debug.append(f"[警告] 列名映射不全，实际映射: {col_map}，尝试按位置提取")
            if len(df.columns) >= 6:
                df_out = df.iloc[:, :6].copy()
                df_out.columns = required
            else:
                return pd.DataFrame(), latest_file, debug
        else:
            df_out = df[[col_map[req] for req in required]].copy()
            df_out.columns = required

        # ══════════════════════════════════════
        # 关键修复：三级城市名填充策略
        # ══════════════════════════════════════
        df_out['省份'] = df_out['省份'].fillna('').astype(str).str.strip()
        df_out['城市'] = df_out['城市'].fillna('').astype(str).str.strip()

        extract_count = 0
        province_fill_count = 0
        unknown_count = 0

        def _fill_city_smart(row):
            """对单行进行智能城市名填充"""
            nonlocal extract_count, province_fill_count, unknown_count

            raw_city = row['城市']
            province = row['省份']

            # 第1级：已有有效城市名 → 直接使用
            if raw_city and raw_city.strip() and raw_city not in ['nan', 'None', '—']:
                return raw_city.strip()

            # 第2级：【核心】从该行的所有文本列中提取城市名
            # 遍历该行所有列（包括非标准列），拼接后搜索
            full_text = ''
            for col in df_out.columns:
                val = row.get(col, '')
                if pd.notna(val) and str(val).strip():
                    full_text += str(val).strip()
            # 也遍历不在标准列中的原始列
            for col in df.columns:
                if col not in required:
                    val = row.get(col, '') if col in row.index else ''
                    if pd.notna(val) and str(val).strip() and len(str(val)) > 3:
                        full_text += str(val).strip()

            extracted = _extract_city_from_text(full_text)
            if extracted:
                extract_count += 1
                return extracted

            # 第3级：用省份名 + "（省级预警）" 兜底
            if province and province.strip() and province not in ['nan', 'None', '', '未知地区']:
                province_fill_count += 1
                return f"{province}（省级预警）"

            # 第4级：完全无法识别
            unknown_count += 1
            return "未知地区"

        df_out['城市'] = df_out.apply(_fill_city_smart, axis=1)

        # 清理省份列的空值
        df_out['省份'] = df_out['省份'].replace('', '未知地区')

        # 过滤掉完全无意义的行
        df_out = df_out.dropna(subset=['预警类型', '预警等级'], how='all')

        debug.append(f"[填充统计] 从文本提取={extract_count}, 省级兜底={province_fill_count}, 未知={unknown_count}")
        debug.append(f"[完成] 有效记录: {len(df_out)} 条")

        # 输出几条样例供验证
        sample_cities = df_out['城市'].head(20).tolist()
        debug.append(f"[样例] 前20条城市名: {sample_cities}")

        return df_out, latest_file, debug

    except Exception as e:
        debug.append(f"[异常] {str(e)}")
        import traceback
        debug.append(traceback.format_exc())
        return pd.DataFrame(), latest_file, debug
