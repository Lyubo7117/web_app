"""
excel_parser.py
Excel 数据解析工具模块

用于解析 national_aqi_crawler.py 和 weather_alarm_crawler.py 生成的 Excel 文件，
将结果统一转换为标准 DataFrame 格式，供 Streamlit 页面直接使用。
"""

import os
import glob
import pandas as pd


# ==============================
# 标准列名定义
# ==============================
AQI_COLUMNS = [
    'city_name', 'aqi', 'pm25', 'pm10', 'co', 'no2',
    'so2', 'o3', 'pollutant', 'level', 'update_time'
]

ALARM_COLUMNS = [
    'city', 'alarm_type', 'alarm_level', 'alarm_title',
    'publish_time', 'description'
]


# ==============================
# AQI 数据解析
# ==============================
def _find_latest_batch(data_dir: str) -> str:
    """
    在 data_output/aqi/ 目录下，找到最新批次文件夹。
    爬虫按日期创建子文件夹（如 2024-04-22/），本函数返回最新的那个。

    Parameters
    ----------
    data_dir : str
        aqi 数据目录的绝对路径

    Returns
    -------
    str
        最新批次文件夹路径，找不到则返回空字符串
    """
    if not os.path.exists(data_dir):
        return ''

    # 列出所有子目录（按批次/日期存放）
    subdirs = [
        d for d in glob.glob(os.path.join(data_dir, '*'))
        if os.path.isdir(d)
    ]
    if not subdirs:
        return ''

    # 按修改时间排序，取最新的
    subdirs.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return subdirs[0]


def parse_aqi_excel(file_path: str, debug_info=None) -> pd.DataFrame:
    """
    解析爬虫生成的 Excel 文件，返回标准格式 DataFrame
    使用 openpyxl 直接读取固定位置的单元格，绕过 pandas 表头识别问题

    Parameters
    ----------
    file_path : str
        Excel 文件的绝对路径
    debug_info : list, optional
        调试信息列表。传入时将解析日志追加到此列表中。

    Returns
    -------
    pd.DataFrame
        标准化后的 DataFrame
    """
    import pandas as pd
    import os
    from openpyxl import load_workbook

    def _log(msg):
        if debug_info is not None:
            debug_info.append(msg)

    try:
        # 提取城市名
        filename = os.path.basename(file_path)
        city = filename.split('_')[0]

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active

        _log(f"  [解析] {city} — 工作表名={ws.title}，总行数={ws.max_row}")

        records = []
        # 数据从第4行开始，第1列为日期时间
        for row_idx in range(4, ws.max_row + 1):
            cell_a = ws.cell(row_idx, 1).value
            if cell_a is None or '统计' in str(cell_a):
                break

            dt = ws.cell(row_idx, 1).value  # 日期时间
            aqi = ws.cell(row_idx, 4).value  # AQI
            level = ws.cell(row_idx, 5).value  # 等级
            primary = ws.cell(row_idx, 6).value  # 首要污染物
            pm25 = ws.cell(row_idx, 7).value  # PM2.5
            pm10 = ws.cell(row_idx, 8).value  # PM10
            co = ws.cell(row_idx, 9).value  # CO
            no2 = ws.cell(row_idx, 10).value  # NO2
            o3 = ws.cell(row_idx, 11).value  # O3_1h
            so2 = ws.cell(row_idx, 13).value  # SO2（跳过第12列O3_8h）

            records.append({
                'city': city,
                'datetime': dt,
                'aqi': aqi,
                'level': level,
                'primary_pollutant': primary,
                'pm25': pm25,
                'pm10': pm10,
                'co': co,
                'no2': no2,
                'o3': o3,
                'so2': so2
            })

        df = pd.DataFrame(records)
        # 数值列转换
        numeric_cols = ['aqi', 'pm25', 'pm10', 'co', 'no2', 'o3', 'so2']
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')

        _log(f"  [成功] {city} — 行数={len(df)}，实际列名={list(df.columns)}")
        return df

    except Exception as e:
        _log(f"  [失败] {os.path.basename(file_path)} — {type(e).__name__}: {e}")
        return pd.DataFrame()


def get_latest_aqi_snapshot(data_dir: str = None):
    """
    从 data_output/aqi/ 目录获取最新批次的 AQI 快照数据。

    会自动：
    1. 找到最新批次文件夹（按日期时间戳子目录）
    2. 递归遍历该目录下所有 Excel 文件（含区域子文件夹）
    3. 取每个城市的最新一条记录，合并为一个 DataFrame
    4. 为每个城市补充经纬度

    Parameters
    ----------
    data_dir : str, optional
        AQI 数据目录路径。默认为 main/data_output/aqi/

    Returns
    -------
    tuple: (pd.DataFrame, str, list)
        - df: 合并后的 AQI 数据，包含 lat/lon 列
        - run_dir: 最新批次目录路径（用于前端显示）
        - debug_info: 调试信息列表（用于前端展示）
    """
    debug_info = []
    empty_df = pd.DataFrame(columns=AQI_COLUMNS + ['lat', 'lon'])

    try:
        debug_info.append("=" * 50)
        debug_info.append("[开始] get_latest_aqi_snapshot 执行")

        # ---- 解析数据目录 ----
        if data_dir is None:
            base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            data_dir = os.path.join(base_dir, 'data_output', 'aqi')

        debug_info.append(f"[路径] 数据根目录 aqi_root = {data_dir}")
        debug_info.append(f"[路径] 目录是否存在 = {os.path.exists(data_dir)}")

        if not os.path.exists(data_dir):
            debug_info.append("[错误] 数据目录不存在，请确认爬虫已运行")
            return empty_df, '', debug_info

        # ---- 找最新批次 ----
        runs = [d for d in os.listdir(data_dir)
                if os.path.isdir(os.path.join(data_dir, d))]
        debug_info.append(f"[批次] 找到 {len(runs)} 个子目录：{runs}")

        if not runs:
            debug_info.append("[错误] data_output/aqi/ 下没有任何批次子目录")
            return empty_df, '', debug_info

        latest_run = sorted(runs)[-1]
        run_dir = os.path.join(data_dir, latest_run)
        debug_info.append(f"[批次] 最新批次名称 = {latest_run}")
        debug_info.append(f"[路径] run_dir 完整路径 = {run_dir}")

        if not os.path.exists(run_dir):
            debug_info.append(f"[错误] 批次目录路径不存在：{run_dir}")
            return empty_df, run_dir, debug_info

        # ---- 遍历文件 ----
        all_files = []
        for root, dirs, files in os.walk(run_dir):
            for f in files:
                all_files.append(os.path.join(root, f))

        xlsx_files = [f for f in all_files if f.endswith('.xlsx')]
        xlsx_files_filtered = [f for f in xlsx_files if not os.path.basename(f).startswith('全国')]

        debug_info.append(f"[统计] os.walk 遍历到总文件数 = {len(all_files)}")
        debug_info.append(f"[统计] 符合 .xlsx 后缀的文件数 = {len(xlsx_files)}")
        debug_info.append(f"[统计] 过滤掉汇总文件后 = {len(xlsx_files_filtered)}")

        # 列出前10个文件名（避免刷屏）
        for i, fp in enumerate(xlsx_files_filtered[:10]):
            debug_info.append(f"  文件[{i+1}] {os.path.basename(fp)}")
        if len(xlsx_files_filtered) > 10:
            debug_info.append(f"  ... 省略其余 {len(xlsx_files_filtered) - 10} 个文件")

        # ---- 逐个解析 ----
        all_cities_data = []
        success_count = 0
        fail_count = 0

        for file_path in xlsx_files_filtered:
            try:
                df_city = parse_aqi_excel(file_path, debug_info=debug_info)

                if df_city.empty:
                    fail_count += 1
                    continue

                # 取最新一条记录
                if 'datetime' in df_city.columns:
                    row = df_city.sort_values('datetime').iloc[-1]
                else:
                    row = df_city.iloc[-1]

                city_name = str(row.get('city', '未知'))
                aqi_val = row.get('aqi', 'N/A')
                all_cities_data.append(row)
                success_count += 1

            except Exception as e:
                debug_info.append(f"  [失败] {os.path.basename(file_path)} — {type(e).__name__}: {e}")
                fail_count += 1

        debug_info.append(f"[汇总] 成功解析 = {success_count} 个城市，失败/跳过 = {fail_count} 个文件")

        if not all_cities_data:
            debug_info.append(f"[错误] 所有文件解析均无有效数据（run_dir: {run_dir}）")
            debug_info.append("=" * 50)
            return empty_df, run_dir, debug_info

        # ---- 合并 + 补经纬度 ----
        df_snapshot = pd.DataFrame(all_cities_data).reset_index(drop=True)

        # 将 city 列统一重命名为 city_name，保持前端接口一致
        if 'city' in df_snapshot.columns and 'city_name' not in df_snapshot.columns:
            df_snapshot.rename(columns={'city': 'city_name'}, inplace=True)

        try:
            from utils.city_coords import get_lat, get_lon
            df_snapshot['lat'] = df_snapshot['city_name'].apply(get_lat)
            df_snapshot['lon'] = df_snapshot['city_name'].apply(get_lon)
        except Exception as e:
            debug_info.append(f"[警告] 补充经纬度失败：{type(e).__name__}: {e}")
            df_snapshot['lat'] = 0.0
            df_snapshot['lon'] = 0.0

        debug_info.append(f"[完成] 共加载 {len(df_snapshot)} 个城市数据")
        debug_info.append(f"[城市] {', '.join(str(c) for c in df_snapshot['city_name'].tolist())}")
        debug_info.append("=" * 50)

        return df_snapshot, run_dir, debug_info

    except Exception as outer_err:
        debug_info.append(f"[异常] 函数整体捕获到未预期异常：{type(outer_err).__name__}: {outer_err}")
        debug_info.append("=" * 50)
        return empty_df, '', debug_info


def parse_uploaded_excel(uploaded_file) -> pd.DataFrame:
    """
    解析用户通过 Streamlit file_uploader 上传的 Excel 文件。

    自动识别列名并映射为标准格式，同时补充经纬度。

    Parameters
    ----------
    uploaded_file : UploadedFile
        Streamlit 的 file_uploader 返回的对象

    Returns
    -------
    pd.DataFrame
        标准化后的 DataFrame，包含 lat/lon
    """
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl')

        # 列名映射
        col_map = {
            '城市': 'city_name', 'AQI': 'aqi', 'PM2.5': 'pm25',
            'PM10': 'pm10', 'CO': 'co', 'NO2': 'no2', 'SO2': 'so2',
            'O3': 'o3', '首要污染物': 'pollutant', '等级': 'level',
            '数据时间': 'update_time', '监测时间': 'update_time',
        }
        df.rename(columns=col_map, inplace=True)

        # 保留存在的标准列
        existing_cols = [c for c in AQI_COLUMNS if c in df.columns]
        df = df[existing_cols]

        # 补充经纬度
        from utils.city_coords import get_lat, get_lon
        if 'city_name' in df.columns:
            df['lat'] = df['city_name'].apply(get_lat)
            df['lon'] = df['city_name'].apply(get_lon)
        else:
            df['lat'] = 0.0
            df['lon'] = 0.0

        return df

    except Exception as e:
        print(f"[ERROR] 解析上传文件失败：{e}")
        return pd.DataFrame(columns=AQI_COLUMNS + ['lat', 'lon'])


# ==============================
# 气象预警数据解析
# ==============================
def parse_alarm_excel(file_path: str) -> pd.DataFrame:
    """
    解析气象预警 Excel 文件（weather_alarm_crawler 产出）。

    Parameters
    ----------
    file_path : str
        Excel 文件路径

    Returns
    -------
    pd.DataFrame
        标准化后的预警数据
    """
    if not os.path.exists(file_path):
        return pd.DataFrame(columns=ALARM_COLUMNS)

    col_map = {
        '城市': 'city', '预警类型': 'alarm_type',
        '预警等级': 'alarm_level', '预警标题': 'alarm_title',
        '发布时间': 'publish_time', '预警内容': 'description',
    }

    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        df.rename(columns=col_map, inplace=True)

        existing_cols = [c for c in ALARM_COLUMNS if c in df.columns]
        return df[existing_cols]

    except Exception as e:
        print(f"[ERROR] 解析预警 Excel 失败：{e}")
        return pd.DataFrame(columns=ALARM_COLUMNS)
