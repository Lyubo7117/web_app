"""
全国34个省会/直辖市/特别行政区实时AQI小时数据爬虫
数据来源：中国天气网 weather.com.cn
支持：省/自治区/直辖市/特别行政区
接口：d1.weather.com.cn/aqi_all/{city_code}.html

已适配 GitHub Actions 自动化：
  - 输出路径自动指向项目 data_output/aqi/ 目录
  - 同时生成汇总文件供 Streamlit 应用直接读取
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import json
import re
import os
import time
import warnings
import sys
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed

warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════
# 输出路径配置（自动适配本地运行 / GitHub Actions）
# ═══════════════════════════════════════════════

# 无论从哪里运行，都输出到项目根目录下的 data_output/aqi/
PROJECT_ROOT = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)), '..', '..'
))
DATA_OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'data_output', 'aqi')


# ═══════════════════════════════════════════════
# 1. 34城市数据库（含行政区划分类）
# ═══════════════════════════════════════════════

CITIES_BY_REGION = {
    "华北区": [
        {"name": "北京",   "city_code": "101010100", "province": "北京市",   "abbr": "北京"},
        {"name": "天津",   "city_code": "101030100", "province": "天津市",   "abbr": "天津"},
        {"name": "石家庄", "city_code": "101090101", "province": "河北省",   "abbr": "石家庄"},
        {"name": "太原",   "city_code": "101100101", "province": "山西省",   "abbr": "太原"},
        {"name": "呼和浩特","city_code":"101080101", "province": "内蒙古自治区","abbr": "呼和浩特"},
    ],
    "东北区": [
        {"name": "沈阳",   "city_code": "101070101", "province": "辽宁省",   "abbr": "沈阳"},
        {"name": "长春",   "city_code": "101060101", "province": "吉林省",   "abbr": "长春"},
        {"name": "哈尔滨", "city_code": "101050101", "province": "黑龙江省", "abbr": "哈尔滨"},
    ],
    "华东区": [
        {"name": "上海",   "city_code": "101020100", "province": "上海市",   "abbr": "上海"},
        {"name": "南京",   "city_code": "101190101", "province": "江苏省",   "abbr": "南京"},
        {"name": "杭州",   "city_code": "101210101", "province": "浙江省",   "abbr": "杭州"},
        {"name": "合肥",   "city_code": "101220101", "province": "安徽省",   "abbr": "合肥"},
        {"name": "福州",   "city_code": "101230101", "province": "福建省",   "abbr": "福州"},
        {"name": "南昌",   "city_code": "101240101", "province": "江西省",   "abbr": "南昌"},
        {"name": "济南",   "city_code": "101120101", "province": "山东省",   "abbr": "济南"},
    ],
    "华中区": [
        {"name": "郑州",   "city_code": "101180101", "province": "河南省",   "abbr": "郑州"},
        {"name": "武汉",   "city_code": "101200101", "province": "湖北省",   "abbr": "武汉"},
        {"name": "长沙",   "city_code": "101250101", "province": "湖南省",   "abbr": "长沙"},
    ],
    "华南区": [
        {"name": "广州",   "city_code": "101280101", "province": "广东省",   "abbr": "广州"},
        {"name": "南宁",   "city_code": "101300101", "province": "广西壮族自治区","abbr": "南宁"},
        {"name": "海口",   "city_code": "101310101", "province": "海南省",   "abbr": "海口"},
    ],
    "西南区": [
        {"name": "重庆",   "city_code": "101040100", "province": "重庆市",   "abbr": "重庆"},
        {"name": "成都",   "city_code": "101270101", "province": "四川省",   "abbr": "成都"},
        {"name": "贵阳",   "city_code": "101260101", "province": "贵州省",   "abbr": "贵阳"},
        {"name": "昆明",   "city_code": "101290101", "province": "云南省",   "abbr": "昆明"},
        {"name": "拉萨",   "city_code": "101140101", "province": "西藏自治区","abbr": "拉萨"},
    ],
    "西北区": [
        {"name": "西安",   "city_code": "101110101", "province": "陕西省",   "abbr": "西安"},
        {"name": "兰州",   "city_code": "101160101", "province": "甘肃省",   "abbr": "兰州"},
        {"name": "西宁",   "city_code": "101150101", "province": "青海省",   "abbr": "西宁"},
        {"name": "银川",   "city_code": "101170101", "province": "宁夏回族自治区","abbr": "银川"},
        {"name": "乌鲁木齐","city_code":"101130101", "province": "新疆维吾尔自治区","abbr": "乌鲁木齐"},
    ],
    "港澳台区": [
        {"name": "香港",   "city_code": "101320101", "province": "香港特别行政区","abbr": "香港"},
        {"name": "澳门",   "city_code": "101330101", "province": "澳门特别行政区","abbr": "澳门"},
        {"name": "台北",   "city_code": "101340101", "province": "台湾省",   "abbr": "台北"},
    ],
}

# 展平为城市列表
ALL_CITIES = []
for region, cities in CITIES_BY_REGION.items():
    for c in cities:
        c["region"] = region
        ALL_CITIES.append(c)


# ═══════════════════════════════════════════════
# 2. 数据获取
# ═══════════════════════════════════════════════

def fetch_aqi_data(city_code, city_name):
    """从中国天气网获取指定城市24小时AQI小时数据"""
    url = f'https://d1.weather.com.cn/aqi_all/{city_code}.html'
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Referer': 'https://www.weather.com.cn/air/',
        'Accept': '*/*',
        'Accept-Language': 'zh-CN,zh;q=0.9',
    }
    resp = requests.get(url, headers=headers, timeout=20, verify=False)
    resp.encoding = 'utf-8'
    text = resp.text
    json_match = re.search(r'setAirData\((\{.*\})\)', text, re.DOTALL)
    if not json_match:
        raise ValueError(f"[{city_name}] 无法解析数据，响应前200字：{text[:200]}")
    data = json.loads(json_match.group(1))
    return data


# ═══════════════════════════════════════════════
# 3. 数据解析
# ═══════════════════════════════════════════════

def parse_records(raw_data):
    """解析原始数据，生成结构化记录列表"""
    city_name = raw_data.get('name', '')
    station_code = raw_data.get('station', '')
    records_raw = raw_data.get('data', [])
    now = datetime.now()
    records = []

    if records_raw:
        last_hour = int(records_raw[-1]['time'])
        current_hour = now.hour
        base_dt = now.replace(minute=0, second=0, microsecond=0)
        last_dt = base_dt
        if last_hour != current_hour:
            diff = (current_hour - last_hour) % 24
            last_dt = base_dt - timedelta(hours=diff)

        time_map = {}
        for i in range(len(records_raw) - 1, -1, -1):
            hour = int(records_raw[i]['time'])
            time_map[i] = last_dt - timedelta(hours=(len(records_raw) - 1 - i))

    for i, r in enumerate(records_raw):
        dt = time_map.get(i, now - timedelta(hours=len(records_raw)-1-i))

        def safe_val(key, default=0):
            v = r.get(key, '')
            if v == '' or v is None:
                return default
            try:
                return float(v)
            except:
                return default

        record = {
            '日期时间': dt.strftime('%Y-%m-%d %H:00'),
            '日期': dt.strftime('%Y-%m-%d'),
            '时间（小时）': dt.strftime('%H:00'),
            'AQI（空气质量指数）': safe_val('t1'),
            'AQI等级': get_aqi_level(safe_val('t1', 0)),
            '首要污染物': get_dominant_pollutant(r),
            'PM2.5（μg/m³）': safe_val('t3'),
            'PM10（μg/m³）': safe_val('t4'),
            'CO（mg/m³）': safe_val('t5'),
            'NO₂（μg/m³）': safe_val('t6'),
            'O₃_1h均值（μg/m³）': safe_val('t7'),
            'O₃_8h均值（μg/m³）': safe_val('t8'),
            'SO₂（μg/m³）': safe_val('t9'),
            '温度（℃）': safe_val('t10'),
            '湿度（%）': safe_val('t11'),
            '气压（hPa）': safe_val('t12'),
            '降水量（mm）': safe_val('t13'),
            '风向': r.get('t14', '—'),
            '风速（m/s）': safe_val('t15'),
        }
        records.append(record)

    return city_name, station_code, records


def get_aqi_level(aqi):
    try:
        aqi = float(aqi)
    except:
        return '—'
    if aqi <= 50:   return '优'
    elif aqi <= 100: return '良'
    elif aqi <= 150: return '轻度污染'
    elif aqi <= 200: return '中度污染'
    elif aqi <= 300: return '重度污染'
    else:            return '严重污染'


def get_aqi_color(aqi):
    try:
        aqi = float(aqi)
    except:
        return 'FFFFFF'
    if aqi <= 50:   return 'A8E05F'
    elif aqi <= 100: return 'FDD74B'
    elif aqi <= 150: return 'FE9B57'
    elif aqi <= 200: return 'FE6A69'
    elif aqi <= 300: return 'A97ABC'
    else:            return 'A87383'


def get_dominant_pollutant(r):
    try:
        vals = {
            'PM2.5': float(r.get('t3', 0) or 0),
            'PM10':  float(r.get('t4', 0) or 0),
            'O₃':    float(r.get('t7', 0) or 0),
            'NO₂':   float(r.get('t6', 0) or 0),
            'SO₂':   float(r.get('t9', 0) or 0),
            'CO':    float(r.get('t5', 0) or 0) * 10,
        }
        mx = max(vals, key=vals.get)
        return mx if vals[mx] > 0 else '—'
    except:
        return '—'


# ═══════════════════════════════════════════════
# 4. Excel 生成（精美化）
# ═══════════════════════════════════════════════

def create_excel(city_info, records, output_path):
    """生成格式精美、颜色醒目的Excel表格"""
    city_name  = city_info['name']
    province   = city_info['province']
    region     = city_info['region']
    station    = city_info['city_code']

    wb = Workbook()
    ws = wb.active
    ws.title = f'{city_name}AQI小时数据'

    HEADER_BG  = '1B4F8A'
    TITLE_BG   = '0D3B6F'
    ODD_ROW    = 'EBF3FB'
    EVEN_ROW   = 'FFFFFF'
    BORDER_CLR = 'B8CCE4'

    thin = Side(style='thin', color=BORDER_CLR)
    med  = Side(style='medium', color='1B4F8A')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 第1行：大标题 ──
    ws.merge_cells('A1:S1')
    c = ws['A1']
    c.value = f'{city_name}市实时空气质量（AQI）小时数据报表'
    c.font = Font(name='微软雅黑', bold=True, size=16, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40

    # ── 第2行：信息栏 ──
    ws.merge_cells('A2:S2')
    fetch_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c = ws['A2']
    c.value = (f'城市：{city_name}（{province}）|  '
               f'所属区域：{region}  |  '
               f'数据来源：中国天气网 weather.com.cn  |  '
               f'城市代码：{station}  |  '
               f'爬取时间：{fetch_time}')
    c.font = Font(name='微软雅黑', size=9, color='FFFFFF', italic=True)
    c.fill = PatternFill('solid', fgColor=HEADER_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    # ── 第3行：表头 ──
    headers = [
        '日期时间', '日期', '时间\n（小时）',
        'AQI\n指数', 'AQI\n等级', '首要\n污染物',
        'PM2.5\nμg/m³', 'PM10\nμg/m³',
        'CO\nmg/m³', 'NO₂\nμg/m³',
        'O₃_1h\nμg/m³', 'O₃_8h\nμg/m³',
        'SO₂\nμg/m³',
        '温度\n℃', '湿度\n%', '气压\nhPa',
        '降水量\nmm', '风向', '风速\nm/s'
    ]
    header_widths = [18, 12, 9, 8, 10, 10, 9, 9, 8, 9, 10, 10, 9, 8, 8, 10, 10, 10, 9]

    for col, (h, w) in enumerate(zip(headers, header_widths), start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 36

    # ── 数据行 ──
    data_keys = [
        '日期时间', '日期', '时间（小时）',
        'AQI（空气质量指数）', 'AQI等级', '首要污染物',
        'PM2.5（μg/m³）', 'PM10（μg/m³）',
        'CO（mg/m³）', 'NO₂（μg/m³）',
        'O₃_1h均值（μg/m³）', 'O₃_8h均值（μg/m³）',
        'SO₂（μg/m³）',
        '温度（℃）', '湿度（%）', '气压（hPa）',
        '降水量（mm）', '风向', '风速（m/s）'
    ]

    for row_idx, record in enumerate(records, start=4):
        is_odd = (row_idx % 2 == 1)
        row_bg = ODD_ROW if is_odd else EVEN_ROW

        for col_idx, key in enumerate(data_keys, start=1):
            val = record.get(key, '—')
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name='微软雅黑', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border

            if col_idx == 4 and val != '—':
                aqi_color = get_aqi_color(val)
                c.fill = PatternFill('solid', fgColor=aqi_color)
                c.font = Font(name='微软雅黑', size=10, bold=True,
                              color='FFFFFF' if val > 100 else '333333')
            elif col_idx == 5 and val != '—':
                aqi_val = record.get('AQI（空气质量指数）', 0)
                aqi_color = get_aqi_color(aqi_val)
                c.fill = PatternFill('solid', fgColor=aqi_color)
                c.font = Font(name='微软雅黑', size=10, bold=True,
                              color='FFFFFF' if float(aqi_val or 0) > 100 else '333333')
            else:
                c.fill = PatternFill('solid', fgColor=row_bg)

        ws.row_dimensions[row_idx].height = 20

    # ── 统计摘要 ──
    last_data_row = 3 + len(records)
    stat_row = last_data_row + 2

    ws.merge_cells(f'A{stat_row}:S{stat_row}')
    c = ws.cell(row=stat_row, column=1, value='统计摘要')
    c.font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[stat_row].height = 24

    stat_keys  = ['AQI（空气质量指数）', 'PM2.5（μg/m³）', 'PM10（μg/m³）',
                  'NO₂（μg/m³）', 'SO₂（μg/m³）', 'CO（mg/m³）', 'O₃_1h均值（μg/m³）']
    stat_names = ['AQI', 'PM2.5(μg/m³)', 'PM10(μg/m³)', 'NO₂(μg/m³)',
                  'SO₂(μg/m³)', 'CO(mg/m³)', 'O₃_1h(μg/m³)']

    for sr, (sk, sn) in enumerate(zip(stat_keys, stat_names)):
        r = stat_row + 1 + sr
        vals_num = [float(rec[sk]) for rec in records
                    if rec[sk] != '—' and str(rec[sk]).replace('.', '').isdigit()]

        c1 = ws.cell(row=r, column=1, value=sn)
        c1.font = Font(name='微软雅黑', bold=True, size=10)
        c1.fill = PatternFill('solid', fgColor='DCE6F1')
        c1.alignment = Alignment(horizontal='center', vertical='center')
        c1.border = border

        for ci, v in enumerate([max(vals_num), min(vals_num),
                                  round(sum(vals_num)/len(vals_num), 1)], start=2):
            c = ws.cell(row=r, column=ci, value=v)
            c.font = Font(name='微软雅黑', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = border
            if sk == 'AQI（空气质量指数）':
                c.fill = PatternFill('solid', fgColor=get_aqi_color(v))
                c.font = Font(name='微软雅黑', size=10, bold=True,
                              color='FFFFFF' if v > 100 else '333333')
            else:
                c.fill = PatternFill('solid', fgColor='EBF3FB')
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = 'A4'
    wb.save(output_path)


# ═══════════════════════════════════════════════
# 5. 单城市爬取任务
# ═══════════════════════════════════════════════

def crawl_city(city_info, output_dir):
    """爬取单个城市数据并生成Excel，返回状态字典"""
    city_name  = city_info['name']
    city_code  = city_info['city_code']
    region     = city_info['region']

    try:
        raw_data = fetch_aqi_data(city_code, city_name)
        _, station, records = parse_records(raw_data)

        if not records:
            return {"city": city_name, "region": region,
                    "status": "warn", "msg": "无数据"}

        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{city_name}_{ts}.xlsx"
        region_dir = os.path.join(output_dir, region)
        os.makedirs(region_dir, exist_ok=True)
        out_path = os.path.join(region_dir, filename)

        create_excel(city_info, records, out_path)

        latest = records[-1]
        return {
            "city": city_name, "region": region,
            "status": "ok",
            "filename": filename,
            "out_path": out_path,
            "record_count": len(records),
            "time_range": f"{records[0]['日期时间']} ~ {records[-1]['日期时间']}",
            "latest_aqi": f"{latest['AQI（空气质量指数）']}（{latest['AQI等级']}）",
        }

    except Exception as e:
        return {"city": city_name, "region": region,
                "status": "fail", "msg": str(e)}


# ═══════════════════════════════════════════════
# 6. 主程序（并发爬取全部城市）
# ═══════════════════════════════════════════════

def main():
    fetch_ts   = datetime.now().strftime('%Y%m%d_%H%M%S')
    # 输出到 data_output/aqi/{日期}/，每批次一个日期文件夹
    output_dir = os.path.join(DATA_OUTPUT_DIR, fetch_ts)
    os.makedirs(output_dir, exist_ok=True)

    print("=" * 65)
    print("  全国34城AQI实时数据爬虫")
    print("  数据来源：中国天气网 weather.com.cn")
    print("=" * 65)
    print(f"输出目录：{output_dir}\n")

    total = len(ALL_CITIES)
    ok_count   = 0
    fail_cities = []

    # 并发爬取（5线程，兼顾速度与稳定性）
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(crawl_city, c, output_dir): c for c in ALL_CITIES}

        for idx, future in enumerate(as_completed(futures), 1):
            result = future.result()
            status_icon = {"ok": "[OK]", "fail": "[FAIL]", "warn": "[WARN]"}.get(result["status"], "[???]")
            region_pad  = result["region"]

            if result["status"] == "ok":
                ok_count += 1
                print(f"  [{idx:02d}/{total}] {status_icon} {result['city']:<6} | "
                      f"AQI={result['latest_aqi']:<14} | "
                      f"{result['time_range']}")
            else:
                fail_cities.append(result)
                print(f"  [{idx:02d}/{total}] {status_icon} {result['city']:<6} | "
                      f"失败：{result.get('msg','未知错误')[:40]}")

            sys.stdout.flush()

    # ── 汇总报告 ──
    print("\n" + "=" * 65)
    print(f"  爬取完成！成功 {ok_count}/{total} 个城市")
    print(f"  成果目录：{output_dir}")
    print("=" * 65)

    if fail_cities:
        print("\n失败城市列表：")
        for fc in fail_cities:
            print(f"  [FAIL] {fc['city']}（{fc['region']}）：{fc.get('msg','未知错误')[:50]}")

    # 生成汇总Excel
    summary_path = os.path.join(output_dir, f"全国34城AQI汇总_{fetch_ts}.xlsx")
    write_summary(output_dir, fetch_ts, summary_path)
    print(f"\n[SUM] 全局汇总表已生成：{os.path.basename(summary_path)}")

    # 打印目录结构
    print("\n目录结构：")
    for region in sorted(os.listdir(output_dir)):
        region_path = os.path.join(output_dir, region)
        if os.path.isdir(region_path):
            files = sorted(os.listdir(region_path))
            print(f"  [DIR] {region}/")
            for f in files:
                print(f"      [FILE] {f}")


def write_summary(output_dir, fetch_ts, summary_path):
    """生成全国34城AQI汇总表"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = '全国34城AQI汇总'

    HEADER_BG  = '1B4F8A'
    BORDER_CLR = 'B8CCE4'
    thin = Side(style='thin', color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ['区域', '城市', '最新AQI', 'AQI等级', '首要污染物',
               'PM2.5', 'PM10', 'CO', 'NO2', 'O3_1h', 'SO2',
               '温度', '湿度', '风向', '风速', '数据时间']
    widths  = [10, 8, 8, 10, 10, 8, 8, 7, 7, 8, 7, 8, 7, 8, 8, 18]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28

    row = 2
    for region in sorted(os.listdir(output_dir)):
        region_path = os.path.join(output_dir, region)
        if not os.path.isdir(region_path):
            continue
        for filename in sorted(os.listdir(region_path)):
            if not filename.endswith('.xlsx') or '汇总' in filename:
                continue
            city_xlsx = os.path.join(region_path, filename)
            try:
                import openpyxl as ox
                wb2 = ox.load_workbook(city_xlsx, data_only=True)
                ws2 = wb2.active
                last_row = ws2.max_row
                data_row = None
                for r in range(4, last_row + 1):
                    val = ws2.cell(r, 1).value
                    if val and '统计' not in str(val) and '日期时间' not in str(val):
                        data_row = r
                if data_row is None:
                    continue
                aqi_val  = ws2.cell(data_row, 4).value
                level_val = ws2.cell(data_row, 5).value
                dt_val   = ws2.cell(data_row, 1).value
                vals2 = [ws2.cell(data_row, c).value for c in range(6, 16)]
                row_data = [region, filename.split('_')[0],
                            aqi_val, level_val] + vals2[:10] + [dt_val]
                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=row, column=col, value=val)
                    c.font = Font(name='微软雅黑', size=10)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border = border
                    if col == 3 and val and str(val).replace('.','').isdigit():
                        color = get_aqi_color(float(val))
                        c.fill = PatternFill('solid', fgColor=color)
                        c.font = Font(name='微软雅黑', size=10, bold=True,
                                      color='FFFFFF' if float(val) > 100 else '333333')
                    elif col == 4 and val:
                        try:
                            v = float(ws2.cell(data_row, 4).value or 0)
                            color = get_aqi_color(v)
                            c.fill = PatternFill('solid', fgColor=color)
                            c.font = Font(name='微软雅黑', size=10, bold=True,
                                          color='FFFFFF' if v > 100 else '333333')
                        except:
                            pass
                    if col == 1:
                        c.fill = PatternFill('solid', fgColor='EBF3FB')
                        c.font = Font(name='微软雅黑', bold=True, size=10)
                ws.row_dimensions[row].height = 20
                row += 1
            except Exception as e:
                pass

    ws.freeze_panes = 'A2'
    wb.save(summary_path)


if __name__ == '__main__':
    main()
