"""
全国气象预警信息爬虫
数据来源：中央气象台 nmc.cn（http://www.nmc.cn/publish/alarm.html）

接口：http://www.nmc.cn/rest/findAlarm
  - pageNo: 页码
  - pageSize: 每页数量
  - signaltype: 预警类型（空为全部）
  - signallevel: 预警等级（空为全部）
  - province: 省份（空为全部）

已适配 GitHub Actions 自动化：
  - 输出路径自动指向 main/data_output/alarms/ 目录
  - 输出文件名以 "全国预警信息_" 开头，供 alarm_parser.py 识别
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

import requests
import json
import re
import os
import time
import warnings
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ═══════════════════════════════════════════════
# 输出路径配置（自动适配本地运行 / GitHub Actions）
# ═══════════════════════════════════════════════

# 输出到 main/data_output/alarms/
MAIN_DIR = os.path.normpath(os.path.join(
    os.path.dirname(os.path.abspath(__file__)), '..'
))
DATA_OUTPUT_DIR = os.path.join(MAIN_DIR, 'data_output', 'alarms')

# ═══════════════════════════════════════════════
# 预警等级映射
# ═══════════════════════════════════════════════

LEVEL_MAP = {
    '红色': '红色',
    '红色预警': '红色',
    '橙色': '橙色',
    '橙色预警': '橙色',
    '黄色': '黄色',
    '黄色预警': '黄色',
    '蓝色': '蓝色',
    '蓝色预警': '蓝色',
    '白色': '白色',
    '白色预警': '白色',
}

LEVEL_COLOR = {
    '红色': 'FF0000',
    '橙色': 'FF7E00',
    '黄色': 'FFD700',
    '蓝色': '4488FF',
    '白色': 'CCCCCC',
}


def extract_level_from_title(title):
    """从预警标题中提取预警等级"""
    for keyword in ['红色预警信号', '橙色预警信号', '黄色预警信号',
                    '蓝色预警信号', '白色预警信号']:
        if keyword in title:
            return keyword.replace('预警信号', '')
    # 尝试从颜色词匹配
    for level in ['红色', '橙色', '黄色', '蓝色', '白色']:
        if level in title:
            return level
    return '未知'


def extract_alarm_type_from_title(title):
    """从预警标题中提取预警类型（如：暴雨、大风、高温等）"""
    # 常见预警类型关键词
    type_keywords = [
        '台风', '暴雨', '暴雪', '寒潮', '大风', '沙尘暴', '高温',
        '干旱', '雷电', '冰雹', '霜冻', '大雾', '霾', '道路结冰',
        '干旱', '森林火险', '雷雨大风', '强对流', '低温', '雪灾',
        '高温', '臭氧', '海啸', '地质灾害', '山洪灾害', '洪水',
        '渍涝', '大风降温', '强降温', '紫外线', '空气重污染',
        '高温中暑', '干旱', '干热风', '龙卷风', '干雾',
    ]
    for kw in type_keywords:
        if kw in title:
            return kw
    # 尝试从 "发布" 或 "信号" 前面提取
    match = re.search(r'发布(.{2,6}?)(?:预警|信号)', title)
    if match:
        return match.group(1)
    return '其他'


def extract_province_city(title):
    """从预警标题中提取省份和城市"""
    # 标题格式示例：
    # "河南省鹤壁市淇县气象台发布大风蓝色预警信号"
    # "广东省广州市气象台发布暴雨橙色预警信号"
    # "北京市气象台发布高温黄色预警信号"
    # "内蒙古自治区气象台发布寒潮蓝色预警信号"

    # 省份关键词
    provinces = [
        '北京市', '天津市', '上海市', '重庆市',
        '河北省', '山西省', '辽宁省', '吉林省', '黑龙江省',
        '江苏省', '浙江省', '安徽省', '福建省', '江西省', '山东省',
        '河南省', '湖北省', '湖南省', '广东省', '海南省',
        '四川省', '贵州省', '云南省', '陕西省', '甘肃省', '青海省',
        '台湾省',
        '内蒙古自治区', '广西壮族自治区', '西藏自治区',
        '宁夏回族自治区', '新疆维吾尔自治区',
        '香港特别行政区', '澳门特别行政区',
    ]
    # 直辖市短名
    short_provinces = {'北京': '北京市', '天津': '天津市',
                       '上海': '上海市', '重庆': '重庆市'}

    province = ''
    city = ''

    # 先尝试完整省份名
    for prov in provinces:
        if title.startswith(prov) or prov in title[:10]:
            province = prov
            break

    # 短名匹配（直辖市）
    if not province:
        for short, full in short_provinces.items():
            if title.startswith(short):
                province = full
                break

    if province:
        # 提取省份后面的城市名
        # 去掉省份名和"气象台"之间的文字
        remaining = title.replace(province, '').replace(
            short_provinces.get(province[:2], ''), '')
        # 匹配 "XX市" 或 "XX州" 等
        city_match = re.match(r'([\u4e00-\u9fa5]{1,6}?市|[\u4e00-\u9fa5]{1,6}?州|[\u4e00-\u9fa5]{1,6}?区|[\u4e00-\u9fa5]{1,6}?县)', remaining)
        if city_match:
            city = city_match.group(1)

    return province, city


# ═══════════════════════════════════════════════
# 数据获取
# ═══════════════════════════════════════════════

def fetch_alarms(page=1, page_size=50):
    """
    从中央气象台获取预警列表。

    Parameters
    ----------
    page : int
        页码（从1开始）
    page_size : int
        每页数量，默认50

    Returns
    -------
    dict
        API 返回的 JSON 数据
    """
    url = 'http://www.nmc.cn/rest/findAlarm'
    params = {
        'pageNo': page,
        'pageSize': page_size,
        'signaltype': '',
        'signallevel': '',
        'province': '',
        '_': int(time.time() * 1000),  # 防缓存时间戳
    }
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                      '(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
        'Referer': 'http://www.nmc.cn/publish/alarm.html',
        'Accept': 'application/json, text/javascript, */*; q=0.01',
        'Accept-Language': 'zh-CN,zh;q=0.9',
    }

    resp = requests.get(url, params=params, headers=headers, timeout=30)
    resp.encoding = 'utf-8'

    # nmc.cn 返回的可能是 JSONP 格式，需要提取 JSON 部分
    text = resp.text.strip()
    if text.startswith('var') or text.startswith('callback'):
        json_match = re.search(r'\((\{.*\})\)', text, re.DOTALL)
        if json_match:
            text = json_match.group(1)

    return json.loads(text)


def parse_alarm_list(api_data):
    """
    解析 API 返回的预警列表，生成结构化记录列表。

    Returns
    -------
    list[dict]
        预警记录列表
    """
    records = []

    try:
        page_list = api_data['data']['page']['list']
    except (KeyError, TypeError):
        return records

    for item in page_list:
        title = item.get('title', '')
        issue_time = item.get('issuetime', '')
        alert_id = item.get('alertid', '')
        url = item.get('url', '')
        pic = item.get('pic', '')

        province, city = extract_province_city(title)
        level = extract_level_from_title(title)
        alarm_type = extract_alarm_type_from_title(title)

        records.append({
            '省份': province,
            '城市': city,
            '预警类型': alarm_type,
            '预警等级': level,
            '发布时间': issue_time,
            '解除时间': '',
            '预警标题': title,
            '预警ID': alert_id,
            '详情URL': url,
            '图标URL': pic,
        })

    return records


# ═══════════════════════════════════════════════
# Excel 生成
# ═══════════════════════════════════════════════

def create_excel(records, output_path):
    """生成格式精美的预警信息 Excel 表格"""
    wb = Workbook()
    ws = wb.active
    ws.title = '全国气象预警信息'

    HEADER_BG  = '8B0000'
    TITLE_BG   = '5B0000'
    BORDER_CLR = 'D0D0D0'

    thin = Side(style='thin', color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 第1行：大标题 ──
    ws.merge_cells('A1:J1')
    c = ws['A1']
    fetch_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    c.value = f'全国气象预警信息汇总  |  爬取时间：{fetch_time}'
    c.font = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor=TITLE_BG)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # ── 第2行：表头 ──
    headers = ['省份', '城市', '预警类型', '预警等级', '发布时间',
               '解除时间', '预警标题', '预警ID', '详情URL', '图标URL']
    widths = [18, 14, 12, 10, 18, 18, 40, 28, 45, 40]

    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = Font(name='微软雅黑', bold=True, size=10, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor=HEADER_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 28

    # ── 数据行 ──
    for row_idx, record in enumerate(records, start=3):
        for col_idx, key in enumerate(headers, start=1):
            val = record.get(key, '')
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = Font(name='微软雅黑', size=10)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = border

            # 预警等级列着色
            if key == '预警等级' and val in LEVEL_COLOR:
                color = LEVEL_COLOR[val]
                c.fill = PatternFill('solid', fgColor=color)
                c.font = Font(name='微软雅黑', size=10, bold=True,
                              color='FFFFFF' if val in ('红色', '橙色') else '333333')

        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = 'A3'
    wb.save(output_path)


# ═══════════════════════════════════════════════
# 主程序
# ═══════════════════════════════════════════════

def main():
    fetch_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    os.makedirs(DATA_OUTPUT_DIR, exist_ok=True)

    print("=" * 65)
    print("  全国气象预警信息爬虫")
    print("  数据来源：中央气象台 nmc.cn")
    print("=" * 65)
    print(f"输出目录：{DATA_OUTPUT_DIR}\n")

    all_records = []
    page = 1
    max_pages = 10  # 最多爬10页，避免过多请求

    while page <= max_pages:
        try:
            print(f"  正在获取第 {page} 页...")
            api_data = fetch_alarms(page=page, page_size=50)

            records = parse_alarm_list(api_data)
            if not records:
                print(f"  第 {page} 页无数据，停止翻页。")
                break

            all_records.extend(records)
            print(f"  第 {page} 页：获取 {len(records)} 条预警")

            # 检查是否还有下一页
            try:
                total_page = api_data['data']['page']['totalPage']
                if page >= total_page:
                    print(f"  已到达最后一页（共 {total_page} 页）。")
                    break
            except (KeyError, TypeError):
                pass

            page += 1
            time.sleep(0.5)  # 礼貌爬取，避免被封

        except requests.exceptions.Timeout:
            print(f"  第 {page} 页请求超时，跳过。")
            page += 1
            time.sleep(2)
        except Exception as e:
            print(f"  第 {page} 页获取失败：{type(e).__name__}: {e}")
            page += 1
            time.sleep(2)

    if not all_records:
        print("\n[警告] 未能获取到任何预警数据。")
        return

    # 统计
    level_counts = {}
    for r in all_records:
        level = r['预警等级']
        level_counts[level] = level_counts.get(level, 0) + 1

    print(f"\n  ── 爬取完成 ──")
    print(f"  总计：{len(all_records)} 条预警信息")
    print(f"  按等级分布：")
    for level in ['红色', '橙色', '黄色', '蓝色', '白色', '未知']:
        count = level_counts.get(level, 0)
        if count > 0:
            print(f"    {level}：{count} 条")

    # 生成 Excel
    filename = f"全国预警信息_{fetch_ts}.xlsx"
    output_path = os.path.join(DATA_OUTPUT_DIR, filename)
    create_excel(all_records, output_path)
    print(f"\n[OK] Excel 已保存：{output_path}")
    print(f"[OK] 文件名：{filename}")
    print("=" * 65)


if __name__ == '__main__':
    main()
